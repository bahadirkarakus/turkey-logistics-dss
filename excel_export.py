"""
Excel Report Generator — openpyxl
Exports the optimisation result to a formatted .xlsx workbook.
"""

from __future__ import annotations
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── colour palette (matches dark-theme brand) ───────────────────────────────
_NAVY   = "0F172A"
_GREEN  = "BBF7D0"
_GREY   = "E2E8F0"
_LIGHT  = "F1F5F9"
_WHITE  = "FFFFFF"
_AMBER  = "FEF08A"

def _hdr(ws, row: int, col: int, value, width: int | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=_WHITE, size=10)
    cell.fill      = PatternFill("solid", fgColor=_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin()
    if width and ws.column_dimensions[get_column_letter(col)].width < width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell

def _cell(ws, row: int, col: int, value, fill: str | None = None,
          bold: bool = False, align: str = "center", num_fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin()
    if fill:
        cell.fill  = PatternFill("solid", fgColor=fill)
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _thin():
    s = Side(style="thin", color="CBD5E0")
    return Border(left=s, right=s, top=s, bottom=s)

def _title(ws, text: str):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    c = ws.cell(row=1, column=1, value=text)
    c.font      = Font(bold=True, size=13, color=_WHITE)
    c.fill      = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def generate_excel(result: dict, supply: dict, demand: dict,
                   cost: dict, co2: dict, scenario_name: str,
                   saved_scenarios: dict | None = None) -> bytes:
    wb = openpyxl.Workbook()

    # ── 1. Shipment Plan ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Shipment Plan"
    _title(ws, f"Turkey Logistics DSS — {scenario_name}  |  "
               f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    sources    = list(supply.keys())
    warehouses = list(demand.keys())

    # header row
    _hdr(ws, 3, 1, "Source", 16)
    for j, wh in enumerate(warehouses, start=2):
        _hdr(ws, 3, j, wh, 11)
    _hdr(ws, 3, len(warehouses) + 2, "Total Shipped", 13)
    _hdr(ws, 3, len(warehouses) + 3, "Capacity",      12)
    _hdr(ws, 3, len(warehouses) + 4, "Slack",         10)

    # data rows
    for i, src in enumerate(sources, start=4):
        fill = _LIGHT if i % 2 == 0 else _WHITE
        _cell(ws, i, 1, src, fill=fill, bold=True, align="left")
        row_total = 0
        for j, wh in enumerate(warehouses, start=2):
            v = result["shipments"].get((src, wh), 0)
            row_total += v
            _cell(ws, i, j, int(v) if v else None,
                  fill=_GREEN if v > 0 else fill)
        _cell(ws, i, len(warehouses) + 2, int(row_total), fill=fill, bold=True)
        _cell(ws, i, len(warehouses) + 3, supply[src],    fill=fill)
        _cell(ws, i, len(warehouses) + 4, supply[src] - int(row_total), fill=fill)

    # demand footer
    dem_row = len(sources) + 4
    _cell(ws, dem_row, 1, "Demand", fill=_GREY, bold=True, align="left")
    for j, wh in enumerate(warehouses, start=2):
        _cell(ws, dem_row, j, demand[wh], fill=_GREY, bold=True)
    _cell(ws, dem_row, len(warehouses) + 2, sum(demand.values()), fill=_GREY, bold=True)

    ws.freeze_panes = "B4"

    # ── 2. Route Details ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Route Details")
    _title(ws2, "Route Details — Active Routes Only")

    headers = ["Source", "Warehouse", "Units", "Unit Cost (TL)",
               "Total Cost (TL)", "CO₂/unit (kg)", "Total CO₂ (kg)"]
    widths  = [16, 14, 10, 16, 16, 16, 16]
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        _hdr(ws2, 3, j, h, w)

    sorted_routes = sorted(result["shipments"].items(), key=lambda x: -x[1])
    total_cost = total_co2 = 0
    for i, ((src, wh), units) in enumerate(sorted_routes, start=4):
        fill   = _LIGHT if i % 2 == 0 else _WHITE
        uc     = cost[src][wh]
        tc     = round(uc * units, 2)
        co2pu  = co2[src][wh]
        tco2   = round(co2pu * units, 2)
        total_cost += tc
        total_co2  += tco2
        _cell(ws2, i, 1, src,    fill=fill, align="left")
        _cell(ws2, i, 2, wh,     fill=fill, align="left")
        _cell(ws2, i, 3, int(units), fill=fill)
        _cell(ws2, i, 4, uc,     fill=fill, num_fmt="#,##0.00")
        _cell(ws2, i, 5, tc,     fill=fill, num_fmt="#,##0.00")
        _cell(ws2, i, 6, co2pu,  fill=fill, num_fmt="0.000")
        _cell(ws2, i, 7, tco2,   fill=fill, num_fmt="#,##0.00")

    # totals row
    tot_row = len(sorted_routes) + 4
    _cell(ws2, tot_row, 1, "TOTAL", fill=_GREY, bold=True, align="left")
    _cell(ws2, tot_row, 5, round(total_cost, 2),
          fill=_GREY, bold=True, num_fmt="#,##0.00")
    _cell(ws2, tot_row, 7, round(total_co2, 2),
          fill=_AMBER, bold=True, num_fmt="#,##0.00")
    for c in [2, 3, 4, 6]:
        _cell(ws2, tot_row, c, None, fill=_GREY)

    # ── 3. Cost & CO₂ Matrix ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Cost Matrix")
    _title(ws3, "Unit Transport Cost Matrix (TL/unit)")
    _hdr(ws3, 3, 1, "Source \\ Warehouse", 18)
    for j, wh in enumerate(warehouses, start=2):
        _hdr(ws3, 3, j, wh, 12)
    for i, src in enumerate(sources, start=4):
        fill = _LIGHT if i % 2 == 0 else _WHITE
        _cell(ws3, i, 1, src, fill=fill, bold=True, align="left")
        for j, wh in enumerate(warehouses, start=2):
            _cell(ws3, i, j, cost[src][wh], fill=fill, num_fmt="#,##0.00")

    ws4 = wb.create_sheet("CO₂ Matrix")
    _title(ws4, "CO₂ Emissions per Unit (kg CO₂/unit)")
    _hdr(ws4, 3, 1, "Source \\ Warehouse", 18)
    for j, wh in enumerate(warehouses, start=2):
        _hdr(ws4, 3, j, wh, 12)
    for i, src in enumerate(sources, start=4):
        fill = _LIGHT if i % 2 == 0 else _WHITE
        _cell(ws4, i, 1, src, fill=fill, bold=True, align="left")
        for j, wh in enumerate(warehouses, start=2):
            _cell(ws4, i, j, co2[src][wh], fill=fill, num_fmt="0.000")

    # ── 4. Scenario Comparison (if available) ───────────────────────────────
    if saved_scenarios and len(saved_scenarios) > 1:
        ws5  = wb.create_sheet("Scenario Comparison")
        _title(ws5, "Scenario Comparison")
        hdrs = ["Scenario", "Total Cost (TL)", "Delta (TL)", "Delta (%)"]
        wids = [24, 18, 14, 12]
        for j, (h, w) in enumerate(zip(hdrs, wids), start=1):
            _hdr(ws5, 3, j, h, w)
        base = list(saved_scenarios.values())[0]["total_cost"]
        for i, (sname, sdata) in enumerate(saved_scenarios.items(), start=4):
            tc   = sdata["total_cost"]
            diff = tc - base
            pct  = diff / base * 100 if base else 0
            fill = _LIGHT if i % 2 == 0 else _WHITE
            _cell(ws5, i, 1, sname,          fill=fill, align="left")
            _cell(ws5, i, 2, tc,             fill=fill, num_fmt="#,##0.00")
            _cell(ws5, i, 3, round(diff, 2), fill=fill, num_fmt="+#,##0.00;-#,##0.00;0")
            _cell(ws5, i, 4, round(pct, 2),  fill=fill, num_fmt='+0.00"%";-0.00"%"')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
