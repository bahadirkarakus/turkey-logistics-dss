"""Tests for PDF and Excel export modules."""

import io

import openpyxl
import pytest

from data import compute_co2_matrix, get_scenario_data
from excel_export import generate_excel
from model import solve
from report import generate_pdf


@pytest.fixture(scope="module")
def result_data():
    supply, demand, cost = get_scenario_data("Normal Season")
    result = solve(supply, demand, cost)
    co2    = compute_co2_matrix()
    return result, supply, demand, cost, co2


class TestPDFExport:
    def test_returns_bytes(self, result_data):
        result, supply, demand, cost, _ = result_data
        pdf = generate_pdf(result, supply, demand, cost, "Normal Season")
        assert isinstance(pdf, bytes)

    def test_minimum_size(self, result_data):
        result, supply, demand, cost, _ = result_data
        pdf = generate_pdf(result, supply, demand, cost, "Normal Season")
        assert len(pdf) > 2_000, "PDF suspiciously small"

    def test_pdf_header_magic(self, result_data):
        result, supply, demand, cost, _ = result_data
        pdf = generate_pdf(result, supply, demand, cost, "Normal Season")
        assert pdf[:4] == b"%PDF", "Output is not a valid PDF"

    def test_with_scenario_comparison(self, result_data):
        result, supply, demand, cost, _ = result_data
        saved = {
            "Normal Season":       {"total_cost": result["total_cost"]},
            "Summer Season":       {"total_cost": result["total_cost"] * 1.15},
        }
        pdf = generate_pdf(result, supply, demand, cost, "Normal Season",
                           saved_scenarios=saved)
        assert len(pdf) > 2_000

    def test_with_monte_carlo(self, result_data):
        result, supply, demand, cost, _ = result_data
        mc = {
            "n_simulations": 50, "mean_cost": 390_000,
            "std_cost": 15_000, "p5_cost": 365_000,
            "p95_cost": 420_000, "n_infeasible": 0,
        }
        pdf = generate_pdf(result, supply, demand, cost, "Normal Season",
                           mc_result=mc)
        assert pdf[:4] == b"%PDF"


class TestExcelExport:
    def test_returns_bytes(self, result_data):
        result, supply, demand, cost, co2 = result_data
        xlsx = generate_excel(result, supply, demand, cost, co2, "Normal Season")
        assert isinstance(xlsx, bytes)

    def test_valid_xlsx(self, result_data):
        result, supply, demand, cost, co2 = result_data
        xlsx = generate_excel(result, supply, demand, cost, co2, "Normal Season")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        assert len(wb.sheetnames) >= 4

    def test_sheet_names(self, result_data):
        result, supply, demand, cost, co2 = result_data
        xlsx = generate_excel(result, supply, demand, cost, co2, "Normal Season")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        assert "Shipment Plan"  in wb.sheetnames
        assert "Route Details"  in wb.sheetnames
        assert "Cost Matrix"    in wb.sheetnames
        assert "CO₂ Matrix"     in wb.sheetnames

    def test_scenario_sheet_added_when_multiple(self, result_data):
        result, supply, demand, cost, co2 = result_data
        saved = {
            "Normal Season":  {"total_cost": 380_000},
            "Summer Season":  {"total_cost": 420_000},
        }
        xlsx = generate_excel(result, supply, demand, cost, co2,
                              "Normal Season", saved_scenarios=saved)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        assert "Scenario Comparison" in wb.sheetnames

    def test_shipment_plan_data(self, result_data):
        result, supply, demand, cost, co2 = result_data
        xlsx = generate_excel(result, supply, demand, cost, co2, "Normal Season")
        wb   = openpyxl.load_workbook(io.BytesIO(xlsx))
        ws   = wb["Shipment Plan"]
        # Row 3 should be the header row
        header = ws.cell(row=3, column=1).value
        assert header is not None
